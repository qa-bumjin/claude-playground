import argparse
import csv
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FILL_HEADER = PatternFill(start_color="E66B2E", end_color="E66B2E", fill_type="solid")
FILL_PASS = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_NA = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FONT_HEADER = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
FONT_DATA = Font(name="맑은 고딕", size=9)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
COL_WIDTHS = {"A": 12, "B": 30, "C": 45, "D": 35, "E": 45, "F": 45, "G": 20, "H": 10, "I": 25}
HEADERS = ["TC ID", "기능", "Rule", "사전 조건", "테스트 단계", "기대 결과", "관련 기획서 페이지", "결과", "Comment"]
CSV_TO_COL = {0: 1, 2: 2, 3: 3, 5: 4, 6: 5, 8: 6, 11: 7}


def build_excel(project_dir: Path, sheet_name: Optional[str]) -> Path:
    output_dir = project_dir / "tc_output"
    csv_files = sorted(output_dir.glob("TC_*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"no TC csv files found in {output_dir}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or project_dir.name[:31]

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    labels = ["전체 TC", "PASS", "FAIL", "N/A", "미실행", "TC 전체 진행률"]
    for ci, label in enumerate(labels, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = Font(name="맑은 고딕", bold=True, size=9)
        c.alignment = ALIGN_C
        c.border = BORDER

    formulas = [
        "=COUNTA(A5:A9999)",
        '=COUNTIF(H5:H9999,"PASS")',
        '=COUNTIF(H5:H9999,"FAIL")',
        '=COUNTIF(H5:H9999,"N/A")',
        "=A2-B2-C2-D2",
        "=IF(A2=0,0,(B2+C2+D2)/A2)",
    ]
    for ci, formula in enumerate(formulas, 1):
        c = ws.cell(row=2, column=ci, value=formula)
        c.font = Font(name="맑은 고딕", size=9)
        c.alignment = ALIGN_C
        c.border = BORDER
        if ci == 6:
            c.number_format = "0.0%"

    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 22
    for ci, header in enumerate(HEADERS, 1):
        c = ws.cell(row=4, column=ci, value=header)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER

    data_row = 5
    for csv_file in csv_files:
        with open(csv_file, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if not row or not row[0].strip():
                    continue
                for csv_idx, col_num in CSV_TO_COL.items():
                    val = row[csv_idx] if csv_idx < len(row) else ""
                    if csv_idx == 6:
                        val = re.sub(r" / (\d+\.)", r"\n\1", val)
                    c = ws.cell(row=data_row, column=col_num, value=val)
                    c.font = FONT_DATA
                    c.border = BORDER
                    c.alignment = ALIGN_C if col_num in (1, 7) else ALIGN_L
                for col_num in (8, 9):
                    c = ws.cell(row=data_row, column=col_num)
                    c.font = FONT_DATA
                    c.border = BORDER
                    c.alignment = ALIGN_C if col_num == 8 else ALIGN_L
                ws.row_dimensions[data_row].height = 45
                data_row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:I4"
    ws.conditional_formatting.add("H5:H9999", CellIsRule(operator="equal", formula=['"PASS"'], fill=FILL_PASS))
    ws.conditional_formatting.add("H5:H9999", CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL))
    ws.conditional_formatting.add("H5:H9999", CellIsRule(operator="equal", formula=['"N/A"'], fill=FILL_NA))

    output_path = project_dir / "TC_마스터.xlsx"
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Export TC CSV files to Excel")
    parser.add_argument("--project", required=True, help="Project directory containing tc_output/")
    parser.add_argument("--sheet-name", default=None, help="Optional Excel sheet name")
    args = parser.parse_args()

    output_path = build_excel(Path(args.project), args.sheet_name)
    print(f"OK exported excel: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

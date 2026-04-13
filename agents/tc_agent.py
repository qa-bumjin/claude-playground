"""
TC 작성 에이전트
PDF 기획서를 청크 단위로 분석 → 기능 인벤토리 생성 → TC CSV 생성 → Excel 마스터 반영
"""
import csv
import json
import os
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

client = anthropic.Anthropic()

# 시스템 프롬프트 로드 (프롬프트 캐싱 대상)
_PROMPT_PATH = Path(__file__).parent.parent / "prompts" / "tc.md"
SYSTEM_PROMPT = _PROMPT_PATH.read_text(encoding="utf-8")

# ── Excel 스타일 상수 (tc.md 스펙 기준) ──────────────────────────────────────
FILL_HEADER = PatternFill(start_color="E66B2E", end_color="E66B2E", fill_type="solid")
FILL_PASS   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FILL_FAIL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_NA     = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FONT_HEADER = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
FONT_DATA   = Font(name="맑은 고딕", size=9)
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
COL_WIDTHS  = {"A": 12, "B": 30, "C": 45, "D": 35, "E": 45, "F": 45, "G": 20, "H": 10, "I": 25}
HEADERS     = ["TC ID", "기능", "Rule", "사전 조건", "테스트 단계", "기대 결과", "관련 기획서 페이지", "결과", "Comment"]
# CSV 열 인덱스 → Excel 열 번호 매핑 (tc.md 컬럼 매핑 기준)
CSV_TO_COL  = {0: 1, 2: 2, 3: 3, 5: 4, 6: 5, 8: 6, 11: 7}


# ── PDF 업로드 ────────────────────────────────────────────────────────────────

def upload_pdf(pdf_path: str) -> str:
    """PDF를 Files API에 업로드하고 file_id를 반환한다."""
    path = Path(pdf_path)
    print(f"  업로드 중: {path.name}")
    with open(path, "rb") as f:
        uploaded = client.beta.files.upload(
            file=(path.name, f, "application/pdf"),
        )
    print(f"  file_id: {uploaded.id}")
    return uploaded.id


# ── 기능 인벤토리 생성 ────────────────────────────────────────────────────────

def build_feature_inventory(file_id: str, total_pages: int) -> dict:
    """
    25페이지 단위(앞뒤 2~3p overlap)로 기획서를 분석해 기능 인벤토리를 생성한다.
    반환값: { feature_id: { feature_name, description, rules, related_pages, status } }
    """
    inventory: dict = {}
    chunk_size = 25
    overlap = 3

    for start in range(1, total_pages + 1, chunk_size - overlap):
        end = min(start + chunk_size - 1, total_pages)
        print(f"  청크 분석: p.{start}~{end}")

        response = client.beta.messages.create(
            model="claude-opus-4-6",
            max_tokens=8192,
            thinking={"type": "adaptive"},
            system=[{
                "type": "text",
                "text": SYSTEM_PROMPT,
                "cache_control": {"type": "ephemeral"},  # 반복 캐시 히트
            }],
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {"type": "file", "file_id": file_id},
                        "title": f"기획서 p.{start}~{end}",
                    },
                    {
                        "type": "text",
                        "text": (
                            f"p.{start}~{end} 범위에서 기능을 추출해 JSON 인벤토리로 반환하세요.\n"
                            f"기존 feature_id가 있으면 신규 항목 생성 없이 정보를 보완만 하세요.\n"
                            f"반환 형식: {{ feature_id: {{feature_name, description, rules, related_pages, status}} }}\n\n"
                            f"기존 인벤토리:\n{json.dumps(inventory, ensure_ascii=False, indent=2)}"
                        ),
                    },
                ],
            }],
            betas=["files-api-2025-04-14"],
        )

        raw = next((b.text for b in response.content if b.type == "text"), "{}")
        # JSON 코드블록 제거
        raw = raw.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        chunk_result = json.loads(raw)
        # 깊은 병합: 기존 feature_id는 보완, 신규는 추가
        for fid, fdata in chunk_result.items():
            if fid in inventory:
                inventory[fid]["rules"].extend(
                    r for r in fdata.get("rules", [])
                    if r.get("rule_id") not in {x["rule_id"] for x in inventory[fid]["rules"]}
                )
                inventory[fid]["related_pages"] = sorted(set(
                    inventory[fid].get("related_pages", []) + fdata.get("related_pages", [])
                ))
                if fdata.get("status") == "complete":
                    inventory[fid]["status"] = "complete"
            else:
                inventory[fid] = fdata

    return inventory


# ── TC 생성 ────────────────────────────────────────────────────────────────────

def generate_tc_for_feature(feature: dict, progress: dict) -> list[dict]:
    """
    기능 1개에 대한 TC를 생성하고 CSV 행 리스트로 반환한다.
    스트리밍으로 출력하며 진행 상태를 progress에 반영한다.
    """
    start_num = progress["total_tc_generated"] + 1
    print(f"\n  TC 생성: {feature.get('feature_name', '')} (TC-{start_num:03d}부터)")

    raw_output = []
    with client.messages.stream(
        model="claude-opus-4-6",
        max_tokens=16000,
        thinking={"type": "adaptive"},
        system=[{
            "type": "text",
            "text": SYSTEM_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }],
        messages=[{
            "role": "user",
            "content": (
                f"아래 기능에 대한 TC를 CSV 형식으로 생성하세요.\n"
                f"TC ID는 TC-{start_num:03d}부터 채번합니다.\n\n"
                f"기능 정보:\n{json.dumps(feature, ensure_ascii=False, indent=2)}\n\n"
                f"CSV 헤더:\n"
                f"TC ID,ASSUMED,기능,Rule,테스트 목적,사전 조건,테스트 단계,입력값,기대 결과,우선순위,분류,관련 기획서 페이지,결과,담당자,Comment"
            ),
        }],
    ) as stream:
        for text in stream.text_stream:
            raw_output.append(text)
            print(text, end="", flush=True)

    full_text = "".join(raw_output)

    # CSV 파싱
    lines = [l for l in full_text.splitlines() if l.strip()]
    tc_rows = []
    in_csv = False
    for line in lines:
        if line.startswith("TC-") or line.startswith('"TC-'):
            in_csv = True
        if in_csv and line.strip():
            tc_rows.append(line)

    parsed = []
    import io
    reader = csv.reader(io.StringIO("\n".join(tc_rows)))
    for row in reader:
        if row and row[0].startswith("TC-"):
            parsed.append(row)

    progress["total_tc_generated"] += len(parsed)
    return parsed


# ── CSV 저장 ──────────────────────────────────────────────────────────────────

def save_tc_csv(feature_id: str, feature_name: str, rows: list[list]) -> str:
    """TC 행 리스트를 CSV 파일로 저장하고 경로를 반환한다."""
    safe_name = "".join(c for c in feature_name if c.isalnum() or c in "_ ")
    filename = f"TC_{feature_id}_{safe_name}.csv"
    out_dir = Path(__file__).parent.parent / "input" / "tc_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    filepath = out_dir / filename

    header = [
        "TC ID", "ASSUMED", "기능", "Rule", "테스트 목적", "사전 조건",
        "테스트 단계", "입력값", "기대 결과", "우선순위", "분류",
        "관련 기획서 페이지", "결과", "담당자", "Comment",
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    print(f"\n  저장: {filepath}")
    return str(filepath)


# ── Excel 마스터 반영 ─────────────────────────────────────────────────────────

def build_excel(csv_files: list[str], sheet_name: str, output_path: str) -> None:
    """
    CSV 파일 목록을 읽어 Excel 마스터 파일을 생성한다.
    tc.md 스펙의 시트 구조·스타일을 그대로 따른다.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # 행 1: 진행률 레이블
    labels = ["전체 TC", "PASS", "FAIL", "N/A", "미실행", "TC 전체 진행률"]
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=1, column=ci, value=lbl)
        c.font = Font(name="맑은 고딕", bold=True, size=9)
        c.alignment = ALIGN_C
        c.border = BORDER

    # 행 2: 진행률 수식
    formulas = [
        "=COUNTA(A5:A9999)",
        '=COUNTIF(H5:H9999,"PASS")',
        '=COUNTIF(H5:H9999,"FAIL")',
        '=COUNTIF(H5:H9999,"N/A")',
        "=A2-B2-C2-D2",
        "=IF(A2=0,0,(B2+C2+D2)/A2)",
    ]
    for ci, f in enumerate(formulas, 1):
        c = ws.cell(row=2, column=ci, value=f)
        c.font = Font(name="맑은 고딕", size=9)
        c.alignment = ALIGN_C
        c.border = BORDER
        if ci == 6:
            c.number_format = "0.0%"

    # 행 3: 빈행
    ws.row_dimensions[3].height = 8

    # 행 4: 헤더
    ws.row_dimensions[4].height = 22
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER

    # 행 5~: 데이터
    data_row = 5
    for csv_file in csv_files:
        with open(csv_file, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)  # 헤더 스킵
            for row in reader:
                if not row or not row[0].strip():
                    continue
                for csv_idx, col_num in CSV_TO_COL.items():
                    val = row[csv_idx] if csv_idx < len(row) else ""
                    if csv_idx == 6:
                        val = val.replace(" / ", "\n")
                    c = ws.cell(row=data_row, column=col_num, value=val)
                    c.font = FONT_DATA
                    c.border = BORDER
                    c.alignment = ALIGN_C if col_num in (1, 7) else ALIGN_L
                for col_num in (8, 9):
                    c = ws.cell(row=data_row, column=col_num)
                    c.font = FONT_DATA
                    c.border = BORDER
                    c.alignment = ALIGN_C if col_num == 8 else ALIGN_L
                ws.row_dimensions[data_row].height = 45
                data_row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:I4"

    # 결과 열 조건부 서식
    ws.conditional_formatting.add("H5:H9999",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=FILL_PASS))
    ws.conditional_formatting.add("H5:H9999",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL))
    ws.conditional_formatting.add("H5:H9999",
        CellIsRule(operator="equal", formula=['"N/A"'],  fill=FILL_NA))

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n  Excel 저장: {output_path}")

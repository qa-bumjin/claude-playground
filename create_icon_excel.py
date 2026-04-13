import csv, re, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

CSV_FILES   = [
    "input/tc_output/TC_ICN001_아이콘리스트검색조건.csv",
    "input/tc_output/TC_ICN002_아이콘리스트결과및수정진입.csv",
]
OUTPUT_FILE = "output/아이콘관리_TC_마스터_테스트.xlsx"
SHEET_NAME  = "아이콘 관리"

FILL_HEADER = PatternFill(start_color="E66B2E", end_color="E66B2E", fill_type="solid")
FILL_PASS   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FILL_FAIL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_NA     = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FONT_HEADER = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
FONT_DATA   = Font(name="맑은 고딕", size=9)
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COL_WIDTHS  = {"A":12,"B":30,"C":45,"D":35,"E":45,"F":45,"G":20,"H":10,"I":25}
HEADERS     = ["TC ID","기능","Rule","사전 조건","테스트 단계","기대 결과","관련 기획서 페이지","결과","Comment"]

# CSV 인덱스 → Excel 열 번호
CSV_TO_COL  = {0:1, 2:2, 3:3, 5:4, 6:5, 8:6, 11:7}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

# 행 1-2: 진행률 집계
for ci, lbl in enumerate(["전체 TC","PASS","FAIL","N/A","미실행","TC 전체 진행률"], 1):
    c = ws.cell(row=1, column=ci, value=lbl)
    c.font = Font(name="맑은 고딕", bold=True, size=9); c.alignment = ALIGN_C; c.border = BORDER
for ci, f in enumerate(["=COUNTA(A5:A9999)",'=COUNTIF(H5:H9999,"PASS")',
                         '=COUNTIF(H5:H9999,"FAIL")', '=COUNTIF(H5:H9999,"N/A")',
                         "=A2-B2-C2-D2", "=IF(A2=0,0,(B2+C2+D2)/A2)"], 1):
    c = ws.cell(row=2, column=ci, value=f)
    c.font = Font(name="맑은 고딕", size=9); c.alignment = ALIGN_C; c.border = BORDER
    if ci == 6: c.number_format = "0.0%"

# 행 3: 빈행
ws.row_dimensions[3].height = 8

# 행 4: 헤더
ws.row_dimensions[4].height = 22
for ci, h in enumerate(HEADERS, 1):
    c = ws.cell(row=4, column=ci, value=h)
    c.fill = FILL_HEADER; c.font = FONT_HEADER; c.alignment = ALIGN_C; c.border = BORDER

# 행 5~: 데이터
data_row = 5
for csv_file in CSV_FILES:
    with open(csv_file, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if not row or not row[0].strip(): continue
            for csv_idx, col_num in CSV_TO_COL.items():
                val = row[csv_idx] if csv_idx < len(row) else ""
                if csv_idx == 6: val = re.sub(r' / (\d+\.)', r'\n\1', val)
                c = ws.cell(row=data_row, column=col_num, value=val)
                c.font = FONT_DATA; c.border = BORDER
                c.alignment = ALIGN_C if col_num in (1, 7) else ALIGN_L
            for col_num in (8, 9):
                c = ws.cell(row=data_row, column=col_num)
                c.font = FONT_DATA; c.border = BORDER
                c.alignment = ALIGN_C if col_num == 8 else ALIGN_L
            rc = ws.cell(row=data_row, column=8)
            rv = str(rc.value or "").strip().upper()
            if rv == "PASS": rc.fill = FILL_PASS
            elif rv == "FAIL": rc.fill = FILL_FAIL
            elif rv == "N/A":  rc.fill = FILL_NA
            ws.row_dimensions[data_row].height = 45
            data_row += 1

ws.freeze_panes = "A5"
ws.auto_filter.ref = "A4:I4"

ws.conditional_formatting.add("H5:H9999",
    CellIsRule(operator="equal", formula=['"PASS"'], fill=FILL_PASS))
ws.conditional_formatting.add("H5:H9999",
    CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL))
ws.conditional_formatting.add("H5:H9999",
    CellIsRule(operator="equal", formula=['"N/A"'],  fill=FILL_NA))

wb.save(OUTPUT_FILE)
print(f"✓ {OUTPUT_FILE} 저장 완료 (TC {data_row - 5}개)")
